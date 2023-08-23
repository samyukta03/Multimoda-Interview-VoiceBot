import speech_recognition as sr
import pyttsx3
import tkinter as tk
import openpyxl

# read the questions from the Excel sheet
wb = openpyxl.load_workbook("D:\FacQA\Answer Analysis\QandAns.xlsx")
ws = wb.active
#questions = [row[0].value for row in ws.iter_rows(min_row=2, values_only=True)]
questions = [cell.value for cell in ws['A'][1:]]

# set up the speech recognition and text-to-speech engines
r = sr.Recognizer()
engine = pyttsx3.init()

# function to ask a question and get the user's response
def ask_question(question):
    # convert the question to speech and speak it
    engine.say(question)
    engine.runAndWait()

    # listen for the user's response
    with sr.Microphone() as source:
        audio = r.listen(source)

    # convert the user's speech to text
    try:
        response = r.recognize_google(audio)
        print("Your response: " + response)
    except sr.UnknownValueError:
        response = "Sorry, I didn't understand that."
        print(response)

    return response

# set up the UI
root = tk.Tk()
root.title("Voice Bot")
root.geometry("400x300")

# add a label to display the questions and responses
label = tk.Label(root, text="", font=("Helvetica", 14))
label.pack(pady=50)

# function to ask the next question and update the label
def ask_next_question():
    if len(questions) > 0:
        # get the next question and ask it
        question = questions.pop(0)
        response = ask_question(question)

        # update the label with the question and response
        label.config(text=question + "\n" + response)
    else:
        # no more questions, end the conversation
        label.config(text="Thank you for talking to me!")
        engine.say("Thank you for talking to me!")
        engine.runAndWait()

# add a button to start the conversation
button = tk.Button(root, text="Start Conversation", command=ask_next_question)
button.pack()

# start the UI
root.mainloop()
