# -*- coding: utf-8 -*-
"""
Created on Wed Mar 22 10:18:33 2023

@author: Dhanya
"""
import torch
from bert_score import score
from transformers import pipeline as p
from sentence_transformers import SentenceTransformer
import nltk
from nltk.translate.bleu_score import sentence_bleu
from nltk.translate.bleu_score import corpus_bleu
import openpyxl
import numpy as np
import pandas as pd
import bert_score.utils as u
def sBert(target,ans):
    if not target or not ans:
        return 0.0  # or return None
    # if ans in ["i don't know", "i am not sure about the answer"]:
    #     return 0.0
    sentence_transformer_model = SentenceTransformer('bert-base-nli-mean-tokens')
    if isinstance(target, str):
       target = [target]
    if isinstance(ans, str):
       ans = [ans]
    # print("yes")
    sentence_embeddings = sentence_transformer_model.encode(target)
    sentence_embeddings2 = sentence_transformer_model.encode(ans)
    # print("yes")
    cos = torch.nn.CosineSimilarity(dim=1, eps=1e-6)
    output = torch.mean(cos(torch.tensor(sentence_embeddings), torch.tensor(sentence_embeddings2))).item()
    return output

def calculate_bleu_score(target, ans):
#     print(target, ans)
#     org=[]
#     org.append(target.split())
#     giv=ans.split()
#    # print(org)
#    # print(giv)
#   #  reference_tokens = nltk.word_tokenize(ans.lower())
#   #  candidate_tokens = nltk.word_tokenize(target.lower())
#     score = sentence_bleu(org,giv)
#     return score
    print(target, ans)
    org=[]
    org.append(target.split())
    giv=ans.split()
   # print(org)
   # print(giv)
   # score = sentence_bleu(org,giv)
    score = sentence_bleu([target], ans, weights= (1,0,0,0))
    return score

def relevance(row_num):
    print(row_num)
    ws= openpyxl.load_workbook('QandAns.xlsx')
    sheet1 = ws['Sheet1']
    row=sheet1.max_row
    target=''
    ans=''
    if sheet1.cell(row=row_num,column=2).value is not None:
     target=sheet1.cell(row=row_num,column=2).value
    if sheet1.cell(row=row_num,column=3).value is not None:
       ans=sheet1.cell(row=row_num,column=3).value
    print(target,ans)
    bert_score=sBert(str(target),str(ans))
    print('Bert score:',bert_score)
    bleu_score = calculate_bleu_score(str(target),str(ans))
    print('Bleu Score',bleu_score)
    if(((bert_score)*100)>50):
       bertscore =  (bert_score*100)
    else:
       bertscore=0
    if((bleu_score*100)>50):
       bleuscore=bleu_score*100
    else:
       bleuscore=0
    print('Bert score:',bertscore)
    print('Bleu score:', bleuscore)
    sheet1.cell(row=row_num,column=4,value=bertscore)
    sheet1.cell(row=row_num,column=5,value=bleuscore)
    ws.save('QandAns.xlsx')
     
relevance(8)
    