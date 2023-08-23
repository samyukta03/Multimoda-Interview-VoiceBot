# -*- coding: utf-8 -*-
"""
Created on Mon Mar 20 14:28:08 2023

@author: Dhanya
"""

import gtts  
from playsound import playsound  
import numpy as np
import os
import matplotlib.pyplot as plt
import pandas as pd
from sklearn.model_selection import train_test_split
import openpyxl
from sklearn.model_selection import train_test_split
import speech_recognition as sr
ws= openpyxl.load_workbook('QandAns.xlsx')
sheet1 = ws['Sheet1']
row=sheet1.max_row
for i in range (1,row):
    os.remove("welcome.mp3")
    text=sheet1.cell(row=i,column=1).value
    t1 = gtts.gTTS(text,lang="en",tld="us")  
    t1.save("welcome.mp3")   
    playsound("welcome.mp3")
    r = sr.Recognizer()
    try:
        with sr.Microphone() as source2:
            r.adjust_for_ambient_noise(source2, duration=0.2)
            audio2 = r.listen(source2)
            MyText = r.recognize_google(audio2)
            MyText = MyText.lower()
            print("Did you say ",MyText)
    except sr.RequestError as e:
        print("Could not request results; {0}".format(e))
         
    except sr.UnknownValueError:
        print("unknown error occurred")
