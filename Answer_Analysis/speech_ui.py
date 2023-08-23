# -*- coding: utf-8 -*-
from flask import Flask,make_response, render_template
import gtts  
from playsound import playsound  
import numpy as np
import os
import pandas as pd
from sklearn.model_selection import train_test_split
import requests
import openpyxl
from sklearn.model_selection import train_test_split
import speech_recognition as sr
import time
import signal
import random
import subprocess
from bs4 import BeautifulSoup
import wave
from datetime import datetime as dt
def signal_handler(sig, frame):
    os._exit(1)


#config = {
#    "DEBUG": True  # run app in debug mode
#}


app = Flask(__name__,template_folder='D:/FacQA/Answer_Analysis')
app.config['TEMPLATES_AUTO_RELOAD'] = True
#app.config.from_mapping(config)

print("Here1")

@app.route('/')
def index():
    return render_template('index_sam.html')
   

@app.route('/run', methods=['POST'])

def run_code():
    print("In here")
    text="Welcome! Lets get you prepared for your interview. Lets start with your behavioural questions"
    t= gtts.gTTS(text,lang="en",tld="us")
    t.save("greet.mp3")
    playsound("greet.mp3")
    os.remove("greet.mp3")
    used_rows_tech = set()
    ws1=openpyxl.load_workbook('D:/Prediction-of-Job-Interview-Performance-by-extracting-Lexical-features-/ques_only.xlsx')
    sheet = ws1['Sheet1']
    row1=sheet.max_row
    results = ''
    for i in  range(1,row1+1):
            #os.remove("welcome.mp3")
            text=sheet.cell(row=i,column=1).value
            print(text)
            t1= gtts.gTTS(text,lang="en",tld="us")  
            t1.save("welcome.mp3")   
            playsound("welcome.mp3")
            os.remove("welcome.mp3")
            r = sr.Recognizer()
            mic = sr.Microphone()
           # time.sleep(20)
            try:
                with mic as source2:
                    print("Inside mic")
                    r.adjust_for_ambient_noise(source2,duration=0.1)
                    # r.energy_threshold = 50000
                    audio2 = r.listen(source2,phrase_time_limit=60,timeout=300)
                    # Save the audio as a WAV file
                    date_string = dt.now().strftime("%d%m%Y%H%M%S")
                    filename = "audio/BehvAns/BehvhAns_" +date_string+ ".wav"
                with wave.open(filename, "wb") as f:
                    f.setnchannels(1)
                    f.setsampwidth(2)
                    f.setframerate(44100)
                    f.writeframes(audio2.get_wav_data())
        
                # Use speech recognition to convert the audio to text
                with sr.AudioFile(filename) as source:
                    audio = r.record(source)
                MyText = r.recognize_google(audio)
                os.remove(filename)
                #MyText = r.recognize_google(audio2)
                MyText = MyText.lower()
                print(MyText)
                if(len(MyText)==0):
                     sheet.cell(row=i,column=2,value=" ")
                else:
                     sheet.cell(row=i,column=2,value=MyText)
                     results += "Did you say " + MyText + "<br>"
                  #  with open(f"audio_{i}.wav", "wb") as f:
                      #  f.write(audio2.get_wav_data())
            except sr.RequestError as e:
                results += "Could not request results; {0}".format(e) + "<br>"
            except sr.UnknownValueError:
                results += "unknown error occurred" + "<br>"
            except KeyboardInterrupt:
                print("Interupted")
                pass
            ws1.save('D:/Prediction-of-Job-Interview-Performance-by-extracting-Lexical-features-/ques_only.xlsx')
    ws1.close()
    import sys
    sys.path.append('D:/Prediction-of-Job-Interview-Performance-by-extracting-Lexical-features-')
    from speech_features import feature_extraction
    from behavioural_lexical_analysis import run_func
    feature_extraction()
    run_func()
    
    text="Okay!Lets move on to your technical round now!"
    t= gtts.gTTS(text,lang="en",tld="us")
    t.save("greet1.mp3")
    playsound("greet1.mp3")
    os.remove("greet1.mp3")
    ws= openpyxl.load_workbook('QandAns.xlsx')
    sheet1 = ws['Sheet1']
    row=sheet1.max_row
    results = ''

    for i in range (1,3): # change to i in the range of 1,10
        row_num = random.randint(1, row)
        while row_num in used_rows_tech:
            row_num = random.randint(1,row)
        used_rows_tech.add(row_num)
        text=sheet1.cell(row=row_num,column=1).value
        print(text)
        t1 = gtts.gTTS(text,lang="en",tld="us")  
        t1.save("welcome.mp3")   
        playsound("welcome.mp3")
        os.remove("welcome.mp3")
        r = sr.Recognizer()
        mic = sr.Microphone()
        time.sleep(20)
        try:
            with mic as source2:
                print("Inside mic")
                r.adjust_for_ambient_noise(source2,duration=1)
                # r.energy_threshold = 50000
                audio2 = r.listen(source2,phrase_time_limit=5)
                # Save the audio as a WAV file
                date_string = dt.now().strftime("%d%m%Y%H%M%S")
                filename = "audio/TechAns/TechhAns_" +date_string+ ".wav"
            with wave.open(filename, "wb") as f:
                f.setnchannels(1)
                f.setsampwidth(2)
                f.setframerate(44100)
                f.writeframes(audio2.get_wav_data())
    
            # Use speech recognition to convert the audio to text
            with sr.AudioFile(filename) as source:
                audio = r.record(source)
            MyText = r.recognize_google(audio)
            os.remove(filename)
            #MyText = r.recognize_google(audio2)
            MyText = MyText.lower()
            print(MyText)
            if(len(MyText)==0):
                 sheet1.cell(row=row_num,column=3,value=" ")
            else:
                 sheet1.cell(row=row_num,column=3,value=MyText)
                 results += "Did you say " + MyText + "<br>"
              #  with open(f"audio_{i}.wav", "wb") as f:
                  #  f.write(audio2.get_wav_data())
        except sr.RequestError as e:
            results += "Could not request results; {0}".format(e) + "<br>"
        except sr.UnknownValueError:
            results += "unknown error occurred" + "<br>"
        except KeyboardInterrupt:
            print("Interupted")
            break
        ws.save('QandAns.xlsx')
        import ans_sim as f1
        f1.relevance(row_num)
        ws =openpyxl.load_workbook('QandAns.xlsx')
        sheet1 = ws['Sheet1']
    ws.close()
    from final_score import compute
    compute(used_rows_tech)
    text="You have successfullyn completed your interview session! Lets look at your feedback"
    t= gtts.gTTS(text,lang="en",tld="us")
    t.save("greet1.mp3")
    playsound("greet1.mp3")
    os.remove("greet1.mp3")
    return "True"

if __name__ == '__main__':
    
 try:
     app.run()
 except Exception as e:
    # print the error message
    print("An error occurred: ", e)
   
print("Here")