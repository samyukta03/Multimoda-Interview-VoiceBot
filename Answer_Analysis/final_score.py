# -*- coding: utf-8 -*-
"""
Created on Sat Apr  1 15:42:54 2023

@author: Dhanya
"""
import openpyxl
import numpy as np
import pandas as pd
import os
def compute(used_rows_tech):

  ws= openpyxl.load_workbook('QandAns.xlsx')
  sheet1 = ws['Sheet1']
  row=sheet1.max_row
  tot_bert=0
  count_bert=0
  tot_bleu=0
  for j in used_rows_tech:
      if sheet1.cell(row=j,column=4).value is not None:
           count_bert=count_bert+1
           print("hey", sheet1.cell(row=j,column=4).value,count_bert)
           tot_bert=tot_bert+ sheet1.cell(row=j,column=4).value
      if sheet1.cell(row=j,column=5).value is not None:
           tot_bleu=tot_bleu+ sheet1.cell(row=j,column=5).value
  if tot_bert!=0:
   tot_bert=tot_bert/count_bert
  if tot_bleu!=0:
   tot_bleu=tot_bleu/count_bert
  print(tot_bert,tot_bleu)
  total_score= (tot_bert+tot_bleu)/2
  print(total_score)
  from openpyxl import Workbook
  #import sys
  #sys.path.append('D:/final_result')
  os.remove('D:/final_result/final_score.xlsx')
  wb = Workbook() 
  ws = wb.create_sheet(title='Final')
  ws.cell(row=1,column=1,value=float(total_score))
  wb.save('D:/final_result/final_score.xlsx')
if __name__ == '__main__':
    compute()
    
