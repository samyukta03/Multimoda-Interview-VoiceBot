# -*- coding: utf-8 -*-
"""
Created on Wed Apr 26 20:33:32 2023

@author: Dhanya
"""
# -*- coding: utf-8 -*-
from flask import Flask,make_response, render_template
import gtts  
from playsound import playsound  
import numpy as np
import os
import pandas as pd
from sklearn.model_selection import train_test_split
import requests
import openpyxl
from sklearn.model_selection import train_test_split
import speech_recognition as sr
import time
import signal
import random
import subprocess
from bs4 import BeautifulSoup
import wave
from datetime import datetime as dt
#import nltk
#nltk.download('punkt')
#nltk.download('averaged_perceptron_tagger')
#nltk.download('maxent_ne_chunker')
#nltk.download('words')
from nltk import sent_tokenize, word_tokenize, pos_tag, ne_chunk

def signal_handler(sig, frame):
    os._exit(1)


#config = {
#    "DEBUG": True  # run app in debug mode
#}


app = Flask(__name__,template_folder='D:/FacQA/Answer_Analysis')
app.config['TEMPLATES_AUTO_RELOAD'] = True
#app.config.from_mapping(config)

print("Here1")

def gen_ques(Text):
    strings= ['I', 'you', 'he', 'she', 'it', 'we', 'they', 'mine', 'yours', 'his', 'hers', 'its', 'ours', 'theirs', 'myself', 'yourself', 'himself', 'herself', 'itself', 'ourselves', 'yourselves', 'themselves', 'this', 'that', 'these', 'those', 'who', 'whom', 'whose', 'what', 'which', 'who', 'whom', 'whose', 'that', 'which', 'anybody', 'anyone', 'anything', 'each', 'either', 'everyone', 'everything', 'neither', 'nobody', 'no one', 'nothing', 'one', 'other', 'somebody', 'someone', 'something']
    s2=['am', 'is', 'are', 'was', 'were', 'been','has', 'have', 'had','do','does', 'do', 'did','go','went','gone','go', 'went','gets', 'get', 'got']
    questions = []
    sentences = sent_tokenize(Text)
    for sentence in sentences:
        # Perform Part-of-Speech tagging on the sentence
        pos_tags = pos_tag(word_tokenize(sentence))
       
        # Identify Named Entities in the sentence
        named_entities = ne_chunk(pos_tags)
       
        # Extract entities, verbs and nouns from the sentence
        entities = [chunk.label() for chunk in named_entities if hasattr(chunk, 'label')]
        verbs = [word for word, tag in pos_tags if tag.startswith('VB')]
        nouns = [word for word, tag in pos_tags if tag.startswith('NN')]
        strings = [noun.lower() for noun in strings]
        s2 = [verb.lower() for verb in s2]
        nouns = [noun.lower() for noun in nouns]
        verbs = [verb.lower() for verb in verbs]
        for item in strings:
            while item in nouns:
              nouns.remove(item)
        for item in s2:
            while item in verbs:
                verbs.remove(item)
        # Generate follow-up questions based on the extracted information
        if entities:
            questions.append(f"What inspired you to become a {entities[0]}?")
        if verbs:
            questions.append(f"Can you tell me more about how you {verbs[0]}?")
        if nouns:
            questions.append(f"What experience do you have with {nouns[0]}?")
    return questions

@app.route('/')
def index():
    return render_template('index_sam.html')
   

@app.route('/run', methods=['POST'])

def run_code():
    print("In here")
    text="Welcome! Lets get you prepared for your interview. Lets start with your behavioural questions"
    t= gtts.gTTS(text,lang="en",tld="us")
    t.save("greet.mp3")
    playsound("greet.mp3")
    os.remove("greet.mp3")
    used_rows_tech = set()
    ws1=openpyxl.load_workbook('D:/Prediction-of-Job-Interview-Performance-by-extracting-Lexical-features-/ques_only.xlsx')
    sheet = ws1['Sheet2']
    row1=sheet.max_row
    files=[]
    results = ''
    for i in  range(1,row1+1):
            #os.remove("welcome.mp3")
            text=sheet.cell(row=i,column=1).value
            print(text)
            t1= gtts.gTTS(text,lang="en",tld="us")  
            t1.save("welcome.mp3")   
            playsound("welcome.mp3")
            os.remove("welcome.mp3")
            r = sr.Recognizer()
            mic = sr.Microphone()
           # time.sleep(20)
            try:
                with mic as source2:
                    print("Inside mic")
                    r.adjust_for_ambient_noise(source2,duration=0.1)
                    # r.energy_threshold = 50000
                    audio2 = r.listen(source2,phrase_time_limit=60,timeout=300)
                    # Save the audio as a WAV file
                    date_string = dt.now().strftime("%d%m%Y%H%M%S")
                    filename = "D:/FacQA/Answer_Analysis/myprosody/myprosody/dataset/audioFiles/BehvhAns_" +date_string+ ".wav"
                    file_for_prosody="BehvhAns_" +date_string+ ".wav"
                    files.append(file_for_prosody)
                with wave.open(filename, "wb") as f:
                    f.setnchannels(1)
                    f.setsampwidth(2)
                    f.setframerate(44100)
                    f.writeframes(audio2.get_wav_data())
                # Use speech recognition to convert the audio to text
                with sr.AudioFile(filename) as source:
                    audio = r.record(source)
                MyText = r.recognize_google(audio)
               # os.remove(filename)
                #MyText = r.recognize_google(audio2)
                MyText = MyText.lower()
                '''
                follow_up=gen_ques(MyText)
                for text in follow_up:
                    print(text)
                    t1= gtts.gTTS(text,lang="en",tld="us")  
                    t1.save("welcome.mp3")   
                    playsound("welcome.mp3")
                    os.remove("welcome.mp3")
                    r1 = sr.Recognizer()
                    mic1 = sr.Microphone()
                   # time.sleep(20)
                    try:
                        with mic1 as source2:
                            print("Inside mic2")
                            r1.adjust_for_ambient_noise(source2,duration=0.1)
                            # r.energy_threshold = 50000
                            audio3 = r.listen(source2,phrase_time_limit=60,timeout=300)
                            # Save the audio as a WAV file
                            date_string = dt.now().strftime("%d%m%Y%H%M%S")
                            filename1 = "D:/FacQA/Answer_Analysis/myprosody/myprosody/dataset/audioFiles/BehvhAns_" +date_string+ ".wav"
                            file_for_prosody="BehvhAns_" +date_string+ ".wav"
                            files.append(file_for_prosody)
                        with wave.open(filename, "wb") as f:
                            f.setnchannels(1)
                            f.setsampwidth(2)
                            f.setframerate(44100)
                            f.writeframes(audio3.get_wav_data())
                
                        # Use speech recognition to convert the audio to text
                        with sr.AudioFile(filename) as source:
                            audio = r.record(source)
                        MyText1 = r.recognize_google(audio)
                        #os.remove(filename)
                        #MyText = r.recognize_google(audio2)
                        MyText1 = MyText1.lower()
                        if(len(MyText1)!=0):
                            MyText=MyText+MyText1
                    except sr.RequestError as e:
                        results += "Could not request results; {0}".format(e) + "<br>"
                    except sr.UnknownValueError:
                        results += "unknown error occurred" + "<br>"
                    except KeyboardInterrupt:
                        print("Interupted")'''
                print(MyText)
                if(len(MyText)==0):
                    sheet.cell(row=i,column=2,value=" ")
                else:
                    sheet.cell(row=i,column=2,value=MyText)
                    results += "Did you say " + MyText + "<br>"
                 #  with open(f"audio_{i}.wav", "wb") as f:
                     #  f.write(audio2.get_wav_data()) 
            except sr.RequestError as e:
                results += "Could not request results; {0}".format(e) + "<br>"
            except sr.UnknownValueError:
                results += "unknown error occurred" + "<br>"
            except KeyboardInterrupt:
                print("Interupted")
                pass
            ws1.save('D:/Prediction-of-Job-Interview-Performance-by-extracting-Lexical-features-/ques_only.xlsx')
    ws1.close()
    import sys
    sys.path.append('D:/Prediction-of-Job-Interview-Performance-by-extracting-Lexical-features-')
    from speech_features import feature_extraction
    from behavioural_lexical_analysis import run_func
    feature_extraction()
    run_func()
    sys.path.append('D:/FacQA/Answer_Analysis/myprosody')
    from testpro import run_code
    
    for name in files:
        run_code(name)
    wb = openpyxl.load_workbook('D:/Prediction-of-Job-Interview-Performance-by-extracting-Lexical-features-/Prosody/MyProsodyFeatures.xlsx')
    ws=wb['Prosody']
    mrow=ws.max_row
    if mrow<=1:
       print("audio not saved")
    else:
        means = []
        count=1
    # Iterate over the columns and calculate the mean of each
        for column in ws.iter_cols(min_row=2):
          column_name = column[0].value
          column_values = [cell.value for cell in column[1:]]
          if(len(column_values)==0):
              continue
          column_mean = sum(column_values) / len(column_values)
          ws.cell(row=2,column=count,value=column_mean)
          count=count+1
          wb.save('D:/Prediction-of-Job-Interview-Performance-by-extracting-Lexical-features-/Prosody/MyProsodyFeatures.xlsx')
        for row in ws.iter_rows(min_row=3):
         ws.delete_rows(row[0].row)
         wb.save('D:/Prediction-of-Job-Interview-Performance-by-extracting-Lexical-features-/Prosody/MyProsodyFeatures.xlsx')
        sys.path.append('D:/Prediction-of-Job-Interview-Performance-by-extracting-Lexical-features-/Prosody')
        from behavioural_prosody_analysis import run_func
        run_func()
    
    text="Okay!Lets move on to your technical round now!"
    t= gtts.gTTS(text,lang="en",tld="us")
    t.save("greet1.mp3")
    playsound("greet1.mp3")
    os.remove("greet1.mp3")
    ws= openpyxl.load_workbook('QandAns.xlsx')
    sheet1 = ws['Sheet1']
    row=sheet1.max_row
    results = ''

    for i in range (1,3): # change to i in the range of 1,10
        row_num = random.randint(1, row)
        while row_num in used_rows_tech:
            row_num = random.randint(1,row)
        used_rows_tech.add(row_num)
        text=sheet1.cell(row=row_num,column=1).value
        print(text)
        t1 = gtts.gTTS(text,lang="en",tld="us")  
        t1.save("welcome.mp3")   
        playsound("welcome.mp3")
        os.remove("welcome.mp3")
        r = sr.Recognizer()
        mic = sr.Microphone()
        time.sleep(5)
        try:
            with mic as source2:
                print("Inside mic")
                r.adjust_for_ambient_noise(source2,duration=1)
                # r.energy_threshold = 50000
                audio2 = r.listen(source2,phrase_time_limit=5)
                # Save the audio as a WAV file
                date_string = dt.now().strftime("%d%m%Y%H%M%S")
                filename = "audio/TechAns/TechhAns_" +date_string+ ".wav"
            with wave.open(filename, "wb") as f:
                f.setnchannels(1)
                f.setsampwidth(2)
                f.setframerate(44100)
                f.writeframes(audio2.get_wav_data())
    
            # Use speech recognition to convert the audio to text
            with sr.AudioFile(filename) as source:
                audio = r.record(source)
            MyText = r.recognize_google(audio)
            os.remove(filename)
            #MyText = r.recognize_google(audio2)
            MyText = MyText.lower()
            print(MyText)
            if(len(MyText)==0):
                 sheet1.cell(row=row_num,column=3,value=" ")
            else:
                 sheet1.cell(row=row_num,column=3,value=MyText)
                 results += "Did you say " + MyText + "<br>"
              #  with open(f"audio_{i}.wav", "wb") as f:
                  #  f.write(audio2.get_wav_data())
        except sr.RequestError as e:
            results += "Could not request results; {0}".format(e) + "<br>"
        except sr.UnknownValueError:
            results += "unknown error occurred" + "<br>"
        except KeyboardInterrupt:
            print("Interupted")
            break
        ws.save('QandAns.xlsx')
        import ans_sim as f1
        f1.relevance(row_num)
        ws =openpyxl.load_workbook('QandAns.xlsx')
        sheet1 = ws['Sheet1']
    ws.close()
    from final_score import compute
    compute(used_rows_tech)
    sys.path.append('D:/Prediction-of-Job-Interview-Performance-by-extracting-Lexical-features-')
    from final_res_mean import run_code
    run_code()
    sys.path.append('D:/Prediction-of-Job-Interview-Performance-by-extracting-Lexical-features-')
    from feedback import run_code
    run_code()
    text="You have successfullyn completed your interview session! Lets look at your feedback"
    t= gtts.gTTS(text,lang="en",tld="us")
    t.save("greet1.mp3")
    playsound("greet1.mp3")
    os.remove("greet1.mp3")
    return "True"

if __name__ == '__main__':
    
 try:
     app.run()
 except Exception as e:
    # print the error message
    print("An error occurred: ", e)
   
print("Here")