# -*- coding: utf-8 -*-
# -*- coding: utf-8 -*-
# -*- coding: utf-8 -*-
"""
Created on Fri Jun 29 16:02:29 2018

@author: sintu
"""

import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from sklearn.model_selection import train_test_split
import openpyxl
from sklearn.model_selection import train_test_split

#Dataset
xx = pd.read_excel('Top_seven_features.xlsx',sheet_name = 'EngagingTone')
x = xx.iloc[:,1:]
ym = pd.read_excel('turker_scores_3.xlsx')
y = ym.iloc[0:,11:12]

print(x,"hey")
print(y,"hey")
#splitting the dataset into training and testing set
x_tr,x_ts,y_tr,y_ts = train_test_split(x,y,test_size=0.1)

#feature scaling
# from sklearn.preprocessing import StandardScaler
# sc_X = StandardScaler()
# sc_y = StandardScaler()
# x_tr = sc_X.fit_transform(x_tr)
# x_ts = sc_X.transform(x_ts)
# y_tr = sc_y.fit_transform(y_tr)
# y_ts = sc_y.fit_transform(y_ts)

#SVR algorithm for training purpose
from sklearn.svm import SVR
regressor = SVR(kernel='rbf')
regressor.fit(x_tr,y_tr)

#SVR algorithm for testing purpose

y_pred = regressor.predict(x_ts)
regressor.score(x_tr,y_tr)
regressor.score(x_ts,y_ts)

#Applying K-Folf cross validation
from sklearn.model_selection import cross_val_score
accuracies = cross_val_score(estimator = regressor, X = x_tr, y = y_tr, cv = 5)
accuracies.mean()
accuracies.std()

#Random forest for training 
from sklearn.ensemble import RandomForestRegressor
regressor1 = RandomForestRegressor(n_estimators = 100, random_state = 0)
regressor1.fit(x_tr,y_tr)

#Random forest for testing purpose and predicting score
print(x_tr, "Training",x_ts,"Testing",y_tr,"training",y_ts,"Testing")
y_pred = regressor.predict(x_ts)
print(y_pred,"hello")
sc1=regressor1.score(x_tr,y_tr)
sc2=regressor1.score(x_ts,y_ts)
#print(sc1,sc2,"Score :")

#Applying K-Folf cross validation on random forest
from sklearn.model_selection import cross_val_score
accuracies1 = cross_val_score(estimator = regressor1, X = x_tr, y = y_tr, cv = 5)
print(accuracies1,"Accuracies")
accuracies1.mean()
accuracies1.std()

#saving score to the sheet
# import openpyxl
ws = openpyxl.load_workbook('Top_seven_features.xlsx')
sheet = ws['EngagingTone']
next_column = sheet.max_column + 1
row_num=64
j=0
for i in range(64,sheet.max_row+1):
    sheet.cell(row=i,column=next_column,value=y_pred[j]*10)
    j=j+1
# sheet.cell(row=141,column=1,value = 'SVR')
# sheet.cell(row=142,column=1,value = 'SVR-->5 values of accuracy by K fold CV')
# for i in range(2,7):
#     sheet.cell(row=142,column=i,value =accuracies[i-2])
# sheet.cell(row=143,column=1,value = 'MeanAccuracies')
# sheet.cell(row=143,column=2,value =accuracies.mean())
# sheet.cell(row=144,column=1,value = 'StdAccuracies')
# sheet.cell(row=144,column=2,value =accuracies.std())
# sheet.cell(row=146,column=1,value = 'Random Forest')
# sheet.cell(row=147,column=1,value = 'RForest-->5 values of accuracy by K fold CV')
# for j in range(2,7):
#     sheet.cell(row=147,column=j,value =accuracies1[j-2])
# sheet.cell(row=148,column=1,value = 'MeanAccuracies')
# sheet.cell(row=148,column=2,value =accuracies1.mean())
# sheet.cell(row=149,column=1,value = 'StdAccuracies')
# sheet.cell(row=149,column=2,value =accuracies1.std())
ws.save('Top_seven_features.xlsx')




# import pandas as pd

# # Load the Excel workbook into a pandas dataframe
# df = pd.read_excel('questions_and_answers.xlsx')

# # Remove duplicate rows based on all columns
# df = df.drop_duplicates()

# # Save the cleaned dataframe back to the Excel workbook
# df.to_excel('questions_and_answers.xlsx', index=False)