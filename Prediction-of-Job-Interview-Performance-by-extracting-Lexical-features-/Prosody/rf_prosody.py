# -*- coding: utf-8 -*-
"""
Created on Thu Apr 27 12:51:23 2023

@author: Dhanya
"""

# -*- coding: utf-8 -*-
"""
Created on Wed Mar 22 12:56:40 2023

@author: Dhanya
"""
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from sklearn.model_selection import train_test_split
import openpyxl
from sklearn.model_selection import train_test_split

xx = pd.read_excel('D:/Prediction-of-Job-Interview-Performance-by-extracting-Lexical-features-/Prosody/Top_seven_features.xlsx',sheet_name = 'EngagingTone')
column_names= xx.columns.tolist()
##print(column_names)
x = xx.iloc[:,0:]
ym = pd.read_excel('D:/Prediction-of-Job-Interview-Performance-by-extracting-Lexical-features-/Prosody/turker_scores_3.xlsx')
y = ym.iloc[0:,11:12]

##print(x,"hey")
##print(y,"hey")
#splitting the dataset into training and testing set

from sklearn.ensemble import RandomForestRegressor
regressor1 = RandomForestRegressor(n_estimators = 100, random_state = 0)
regressor1.fit(x,y)

from sklearn.model_selection import cross_val_score
accuracies1 = cross_val_score(estimator = regressor1, X = x, y = y, cv = 5)
print(accuracies1,"Accuracies")
accuracies1.mean()
accuracies1.std()

df=pd.read_excel('D:/Prediction-of-Job-Interview-Performance-by-extracting-Lexical-features-/Prosody/prosodic_features1.xlsx',sheet_name = 'prosodic_features')
df_subset = df[column_names]

#print(df_subset)
y_pred = regressor1.predict(df_subset)
#print(y_pred*10)
print(y_pred)

from openpyxl import Workbook
wb = Workbook() 
ws = wb.create_sheet(title='Final')
ws.append(['EngagingTone','Calm','Excited','Friendly','NoFillers','Focused','Pause','Speaking rate','Structured answer','Recommended Hiring'])
ws.cell(row=2,column=1,value=(y_pred[0]*10))
wb.save('D:/final_result/final_result_prosody.xlsx')




  