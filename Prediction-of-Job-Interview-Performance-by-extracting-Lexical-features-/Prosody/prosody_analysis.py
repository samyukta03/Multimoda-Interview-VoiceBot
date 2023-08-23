# -*- coding: utf-8 -*-
"""
Created on Tue Apr 25 17:40:11 2023

@author: Dhanya
"""
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from statistics import mean
import csv

def best_fit_slope_and_intercept(xs,ys):
    m = (((mean(xs)*mean(ys)) - mean(xs*ys)) /
         ((mean(xs)*mean(xs)) - mean(xs*xs)))
    b = mean(ys) - m*mean(xs)
    return m, b
def squared_error(ys_orig,ys_line):
	return sum((ys_line - ys_orig) * (ys_line - ys_orig))

def coefficient_of_determination(ys_orig,ys_line):
    y_mean_line = [mean(ys_orig) for y in ys_orig]
    squared_error_regr = squared_error(ys_orig, ys_line)
    squared_error_y_mean = squared_error(ys_orig, y_mean_line)
    return 1 - (squared_error_regr/squared_error_y_mean)

#dataset
x = pd.ExcelFile('prosodic_features1.xlsx')
x1=pd.read_excel(x,'prosodic_features')
y = pd.read_excel('turker_scores_3.xlsx')

import openpyxl
from openpyxl import Workbook
wb = Workbook() 
ws = wb.create_sheet(title='Coefficient of determination')
ws1 = openpyxl.load_workbook('prosodic_features1.xlsx')
ws2 = openpyxl.load_workbook('turker_scores.xlsx')
sheet2 = ws1['prosodic_features']
sheet3 = ws2['turker_scores']

for i in range(1,57):
    ws.cell(row=1,column=i+1,value=(sheet2.cell(row=1,column=i+1).value)) # 1st row 1stcol-> value
    ws.cell(row=i+1,column=1,value=(sheet3.cell(row=1,column=i+1).value)) #
wb.save('Coeff_of_determination.xlsx')

from sklearn.feature_selection import mutual_info_regression
for i in range(1,57):
    filename = "prosodic_features1.csv"
    filename2= "turker_scores_3.csv"
    mi = []
    x2 = []
    with open(filename,'r') as csvfile:
		# creating a csv reader object
        csvreader = csv.reader(csvfile)
        row_num = 0
        for row in csvreader:
            if(row_num<=69):
              x2.append(row[i])
            else:
                break
        row_num += 1
        del(x2[0])
        for j in range(0,len(x2)):
            x2= [float(i) for i in x2]
        #print(x2)
        for m in range(1,19):
            y1 = []
            with open(filename2,'r') as csvfile1:
                csvreader1 = csv.reader(csvfile1)
                row_num = 0
                for row in csvreader1:
                    if(row_num<=69):
                     #print(m)
                     y1.append(row[m])
                    else:
                         break
                row_num += 1
                  # x2[l]=float(x2[l])
                del(y1[0])
                #print(y1)
                for u in range(0,len(y1)):
                    y1= [float(m) for m in y1]
                #print("hey",y1)
                xs = np.array(x2, dtype=np.float64)
                ys = np.array(y1, dtype=np.float64)
            m1, b = best_fit_slope_and_intercept(xs,ys)
            regression_line = [(m1*x)+b for x in xs]
            r_squared = coefficient_of_determination(ys,regression_line)
             #mi = mutual_info_regression(x2,y1)
            # print(mi,"hey",i+1,j+1)
            #print("R sqA",r_squared)
            ws.cell(row=m+1,column=i+1,value=r_squared)
            wb.save('Coeff_of_determination.xlsx')