# -*- coding: utf-8 -*-
"""
Created on Wed Mar 22 12:56:40 2023

@author: Dhanya
"""
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from sklearn.model_selection import train_test_split
import openpyxl
from sklearn.model_selection import train_test_split
import sys

def run_code():
    xx = pd.read_excel('D:/Prediction-of-Job-Interview-Performance-by-extracting-Lexical-features-/Top_seven_features.xlsx',sheet_name = 'EngagingTone')
    column_names= xx.columns.tolist()
    ##print(column_names)
    x = xx.iloc[:,0:]
    ym = pd.read_excel('D:/Prediction-of-Job-Interview-Performance-by-extracting-Lexical-features-/turker_scores_3.xlsx')
    y = ym.iloc[0:,11:12]
    
    #splitting the dataset into training and testing set
    x_tr,x_ts,y_tr,y_ts = train_test_split(x,y,test_size=0.1)
    
    from sklearn.ensemble import RandomForestRegressor
    regressor1 = RandomForestRegressor(n_estimators = 100, random_state = 0)
    regressor1.fit(x,y)
    
    #Random forest for testing purpose and predicting score
    y_pred = regressor1.predict(x_ts)
    regressor1.score(x_tr,y_tr)
    
    print(y_pred[0]*10)
    print(regressor1.score(x_ts,y_ts)*100, " LexicalFeatures")
    sys.path.append('D:/Prediction-of-Job-Interview-Performance-by-extracting-Lexical-features-/Prosody')
    from engaging_rf_prosody import run_code
    run_code(y_pred)
    correlation = np.corrcoef(y_pred, y_ts.squeeze())
    print("Correlation coefficient:", correlation[0, 1])
    wb = openpyxl.load_workbook('coeff.xlsx')
    ws = wb['Coeff']
    ws.cell(row=3,column=1,value='Engaging')
    ws.cell(row=3,column=2,value=correlation[0,1])
    wb.save('coeff.xlsx')
    
    
    from sklearn.model_selection import cross_val_score
    accuracies1 = cross_val_score(estimator = regressor1, X = x, y = y, cv = 5)
    accuracies1.mean()
    accuracies1.std()
    
    df=pd.read_excel('D:/Prediction-of-Job-Interview-Performance-by-extracting-Lexical-features-/Input_Features.xlsx',sheet_name = 'Features')
    df_subset = df[column_names]
    
    y_pred = regressor1.predict(df_subset)
      
    from openpyxl import Workbook
    wb = Workbook() 
    ws = wb.create_sheet(title='Final')
    ws.append(['EngagingTone','Calm','Excited','Friendly','NoFillers','Focused','Pause','SpeakingRate','StructuredAnswers','RecommendedHiring'])
    ws.cell(row=2,column=1,value=(y_pred[0]*10))
    wb.save('D:/final_result/final_result.xlsx')
    

plt.show()

run_code()
    
    
    
      