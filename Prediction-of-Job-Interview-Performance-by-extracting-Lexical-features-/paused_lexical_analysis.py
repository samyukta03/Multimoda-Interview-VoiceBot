import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from sklearn.model_selection import train_test_split
import openpyxl
from sklearn.model_selection import train_test_split

def run_code():
    xx = pd.read_excel('D:/Prediction-of-Job-Interview-Performance-by-extracting-Lexical-features-/Top_seven_features.xlsx',sheet_name = 'Paused')
    column_names= xx.columns.tolist()
    # print(column_names)
    x = xx.iloc[:,0:]
    ym = pd.read_excel('D:/Prediction-of-Job-Interview-Performance-by-extracting-Lexical-features-/turker_scores_3.xlsx')
    y = ym.iloc[0:,10:11]
    x_tr,x_ts,y_tr,y_ts = train_test_split(x,y,test_size=0.1)
    # print(x,"hey")
    # print(y,"hey")
    #splitting the dataset into training and testing set
    
    from sklearn.ensemble import RandomForestRegressor
    regressor1 = RandomForestRegressor(n_estimators = 100, random_state = 0)
    regressor1.fit(x,y)
    y_pred = regressor1.predict(x_ts)
    regressor1.score(x_tr,y_tr)
    
    print(regressor1.score(x_ts,y_ts)*100)
    correlation = np.corrcoef(y_pred, y_ts.squeeze())
    print("Correlation coefficient:", correlation[0, 1])
    wb = openpyxl.load_workbook('coeff.xlsx')
    ws = wb['Coeff']
    ws.cell(row=7,column=1,value='Pause')
    ws.cell(row=7,column=2,value=correlation[0,1])
    wb.save('coeff.xlsx')
    from sklearn.model_selection import cross_val_score
    accuracies1 = cross_val_score(estimator = regressor1, X = x, y = y, cv = 5)
    # print("\nAccuracies\n",accuracies1)
    accuracies1.mean()
    accuracies1.std()
    
    # df=pd.read_excel('Input_Features.xlsx',sheet_name = 'Features') #-for tech qns features extracted
    df=pd.read_excel('D:/Prediction-of-Job-Interview-Performance-by-extracting-Lexical-features-/Input_Features.xlsx',sheet_name = 'Features')
    df_subset = df[column_names]
    
    # print(df_subset)
    y_pred = regressor1.predict(df_subset)
    # print(y_pred*10)
    
    wb = openpyxl.load_workbook('D:/final_result/final_result.xlsx')
    ws=wb['Final']
    ws.cell(row=2,column=7,value=(y_pred[0]*10))
    wb.save('D:/final_result/final_result.xlsx')
run_code()