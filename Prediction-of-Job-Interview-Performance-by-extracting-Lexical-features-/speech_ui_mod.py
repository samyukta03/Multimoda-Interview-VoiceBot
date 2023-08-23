# -*- coding: utf-8 -*-
"""
Created on Wed Mar 22 16:56:33 2023

@author: Dhanya
"""

# -*- coding: utf-8 -*-
from flask import Flask, render_template, request
import gtts  
from playsound import playsound  
import numpy as np
import os
import matplotlib.pyplot as plt
import pandas as pd
from sklearn.model_selection import train_test_split
import openpyxl
from sklearn.model_selection import train_test_split
import speech_recognition as sr
import time

app = Flask(__name__,template_folder='D:/FacQA/Answer Analysis')

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/run', methods=['POST'])
def run_code():
    ws= openpyxl.load_workbook('QandAns.xlsx')
    sheet1 = ws['Sheet1']
    row=sheet1.max_row
    results = ''
    try:
     for i in range (1,row):
        os.remove("welcome.mp3")
        text=sheet1.cell(row=i,column=1).value
        t1 = gtts.gTTS(text)  
        t1.save("welcome.mp3")   
        playsound("welcome.mp3")
        r = sr.Recognizer()
        mic = sr.Microphone()
        time.sleep(20)
        try:
            with mic as source2:
                print("Inside mic")
               # r.adjust_for_ambient_noise(source2)
               # r.energy_threshold = 50000
                audio2 = r.listen(source2,phrase_time_limit=5)
                MyText = r.recognize_google(audio2)
                MyText = MyText.lower()
                print(MyText)
                sheet1.cell(row=i,column=3,value=MyText)
                results += "Did you say " + MyText + "<br>"
              #  with open(f"audio_{i}.wav", "wb") as f:
                  #  f.write(audio2.get_wav_data())
        except sr.RequestError as e:
            results += "Could not request results; {0}".format(e) + "<br>"
        except sr.UnknownValueError:
            results += "unknown error occurred" + "<br>"
        ws.save('QandAns.xlsx')
    finally:
        r.__exit__(None, None, None)
        
    ws.close()
    return results

if __name__ == '__main__':
    app.run()