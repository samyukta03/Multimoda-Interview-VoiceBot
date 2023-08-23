# -*- coding: utf-8 -*-
"""
Created on Fri Apr 28 00:45:27 2023

@author: SAMYU
"""

# Importing the libraries
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import openpyxl

def run_code():
    ws1 = openpyxl.load_workbook('D:/final_result/final_result.xlsx')
    sheet1 = ws1['Final']
    
    ws2 = openpyxl.load_workbook('D:/final_result/myprosody_final.xlsx')
    sheet2 = ws2['Final']
    
    list1=[]
    list2=[]
    
    for i in range(1,9):
        list1.append(sheet1.cell(row=2,column=i).value)
        list2.append(sheet2.cell(row=2,column=i).value)
    
    print(list1,list2)
    
    list3 = []
    for i in range(len(list1)):
        list3.append((list1[i]+list2[i])/2)
    print(list3)
    
    for j in range(len(list3)):
        sheet1.cell(row=2, column=j+1, value=list3[j])
    
    # Save the result
    ws1.save('D:/final_result/final_result.xlsx')
