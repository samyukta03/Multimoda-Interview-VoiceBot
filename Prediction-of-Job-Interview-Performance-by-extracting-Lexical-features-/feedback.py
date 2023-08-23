# -*- coding: utf-8 -*-
"""
Created on Sat Apr 29 20:04:45 2023

@author: Dhanya
"""

import numpy as np
import pandas as pd
import openpyxl

def run_code():
    ws1 = openpyxl.load_workbook('D:/final_result/final_result.xlsx')
    sheet1 = ws1['Final']
    
    
    mrow= sheet1.max_row
    mcol=sheet1.max_column
    
    print(mrow,mcol)
    
    for i in range(1,10):
        val=sheet1.cell(row=2,column=i).value
        if val<50 :
            ws2= openpyxl.load_workbook('D:/Prediction-of-Job-Interview-Performance-by-extracting-Lexical-features-/Top_seven_features.xlsx')
            attr=sheet1.cell(row=1,column=i).value
            sheet2=ws2[attr]
            column_names=[]
            for cell in sheet2[1]:
               column_names.append(cell.value)
            print(column_names)
            seen=set()
            feedback='Feedback : '
            arr=['Anxiety', 'Cognitive', 'Perceptual', 'FillerWords', 'PosEmotion', 'NegEmotion', 'Confidence', 'Sadness']
            for att in column_names:
                if att in arr and att not in seen:
                    if att=='Anxiety':
                       feedback=feedback+' You need not be so tensed in the interview.Being more prepared and with more practice you will be ready!.'
                    if att=='Cognitive':
                       feedback=feedback+ ' You can improve your Cognitive skills which are the abilities that allow us to learn, process, analyze, and retain information.Practice more problem solving exercises.'
                    if att=='Perceptual':
                        feedback=feedback+ ' Pay attention to details and practice active listening in order to improve your percetual skills.'
                    if att=='FillerWords':
                       feedback=feedback+ ' Slow down your rate of speaking to avoid using a lot filler words and give yourself time to think about what you want to say.'
                    if att=='PosEmotion':
                       feedback= feedback+ ' Show interest using positive language that convey your enthusiasm in the company and for the role.'
                    if att=='NegEmotion':
                        feedback= feedback+' Language has significant impact so be mindful of the tone and body language also avoid using absolute words like "always," "never," "completely," or "totally". '
                    if att=='Confidence':
                        feedback= feedback+' Do your reseach, ask more questions and be prepared to feel more confident.'
                    if att=='Sadness':
                        feedback= feedback+' Focus on your strengths and Visualize success to feel more positive.'
                    seen.add(att)
            sheet1.cell(row=3,column=i,value=feedback)
            ws1.save('D:/final_result/final_result.xlsx')
        else:
            sheet1.cell(row=3,column=i,value=' Keep Practicing!! You are good to go!')
            ws1.save('D:/final_result/final_result.xlsx')
            