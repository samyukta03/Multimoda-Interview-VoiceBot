# -*- coding: utf-8 -*-
"""
Spyder Editor
Created on Mon June  20 15:41:12 2018

@author: sintu

"""
# Importing the libraries
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd

#dataset
x = pd.ExcelFile('features2.xlsx')
x1=pd.read_excel(x,'Feature')
y = pd.read_excel('turker_scores_3.xlsx')
# print(x1.head(2),y.head(2))
#Creating Mutual Information file
import openpyxl
from openpyxl import Workbook
wb = Workbook() 
ws = wb.create_sheet(title='Mutual Information')
ws1 = openpyxl.load_workbook('features.xlsx')
ws2 = openpyxl.load_workbook('turker_scores.xlsx')
sheet2 = ws1['Feature']
sheet3 = ws2['turker_scores']

for i in range(1,23):
    ws.cell(row=1,column=i+1,value=(sheet2.cell(row=1,column=i+1).value)) # 1st row 1stcol-> value
    ws.cell(row=i+1,column=1,value=(sheet3.cell(row=1,column=i+1).value)) #
wb.save('MI_score.xlsx')

# Calculating Mutual Information
from sklearn.feature_selection import mutual_info_regression
for i in range(1,23):
    mi = []
    y1 = []
    x2 = x1.iloc[0:,i-1:i]
    for j in range(1,20):
       y1 = y.iloc[0:,j-1:j]
#     print("hey1",y1,"hey",i)
#     print("x value ",x)
       print(x2)
       print(y1)
       mi = mutual_info_regression(x2,y1)
       print(mi,"hey",i+1,j+1)
       ws.cell(row=j+1,column=i+1,value=mi[0])
       wb.save('MI_score.xlsx')
 
                 #col     col2   col3 
        #row1
        
        #row2
        #row3
        #row4
