# -*- coding: utf-8 -*-
"""
Created on Tue May 23 21:39:45 2023

@author: Dhanya
"""

import numpy as np
import matplotlib.pyplot as plt
import pandas as pd

#dataset
x = pd.ExcelFile('Top_seven_features.xlsx')
x1=pd.read_excel(x,'Friendly')
y = pd.read_excel('turker_scores_3.xlsx')
# print(x1.head(2),y.head(2))

import openpyxl
from openpyxl import Workbook
wb = Workbook() 
ws = wb.create_sheet(title='Coeff of determination')
ws1 = openpyxl.load_workbook('Top_seven_features.xlsx')
ws2 = openpyxl.load_workbook('turker_scores.xlsx')
sheet2 = ws1['Friendly']
sheet3 = ws2['turker_scores']

for i in range(1,9):
    ws.cell(row=1,column=i+1,value=(sheet2.cell(row=1,column=i).value)) # 1st row 1stcol-> value
    ws.cell(row=i+1,column=1,value=(sheet3.cell(row=1,column=i+1).value)) #
wb.save('coeff.xlsx')
