import openpyxl
import numpy as np
import pandas as pd
from openpyxl import Workbook
wb = Workbook() 
ws = wb.create_sheet(title='Scores')
ws2 = openpyxl.load_workbook('scores_temp.xlsx')
sheetx=ws2['scores']
c = sheetx.max_column
for i in range(1,36):
     for j in range(0,c):
          ws.cell(row=i+1,column=j+1,value=sheetx.cell(row=i+1,column=j+1))
start_row = 36
every_nth_row = 10
selected_rows = sheetx.iloc[start_row-1::every_nth_row]

# Export the selected rows to a new Excel file
selected_rows.to_excel(ws, index=False)